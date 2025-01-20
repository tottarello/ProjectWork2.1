import sqlite3
from openpyxl import load_workbook

file_exists = False  # Flag per il file utenti.xlsx

try:
    # Connessione al database SQLite
    try:
        conn = sqlite3.connect("utenti.db")
        cursor = conn.cursor()
    except sqlite3.Error as e:
        print("Errore nella connessione al database: " + str(e))

    try:
        # Creazione della tabella SQL
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS utenti (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT,
            cognome TEXT,
            email TEXT,
            telefono TEXT
        )
        """)
    except sqlite3.Error as e:
        print("Errore nella creazione della tabella: " + str(e))

    try:
        # Caricamento del file Excel
        try:
            wb = load_workbook("utenti.xlsx")
            ws = wb.active
            file_exists = True  # Setto a True il Flag
        except FileNotFoundError as e:
            print("Errore: il file 'utenti.xlsx' non esiste.")
        except Exception as e:
            print("Errore durante il caricamento del file Excel: " + str(e))

        # Inserimento dei dati dalla seconda riga (saltando le intestazioni)
        if file_exists:  # Se esiste porcedo
            try:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    cursor.execute("""
                    INSERT INTO utenti (nome, cognome, email, telefono)
                    VALUES (?, ?, ?, ?)
                    """, row)
            except sqlite3.Error as e:
                print("Errore durante l'inserimento dei dati: " + str(e))

    except Exception as e:
        print("Errore durante la gestione del file Excel o dei dati: " + str(e))

    if file_exists:  # fa il commit solo se esiste
        try:
            # Commit delle modifiche
            conn.commit()
            print("Dati salvati nel database con successo.")
        except sqlite3.Error as e:
            print("Errore durante il commit: " + str(e))

finally:
    try:
        # Chiusura della connessione
        if conn:
            conn.close()
            print("Connessione al database chiusa.")
    except sqlite3.Error as e:
        print("Errore durante la chiusura della connessione: " + str(e))
